from __future__ import annotations

from pathlib import Path


def extract_docx(docx_path: Path, *, max_lines: int = 350) -> None:
    from docx import Document  # type: ignore

    doc = Document(str(docx_path))
    lines: list[str] = []
    for p in doc.paragraphs:
        t = (p.text or "").strip()
        if t:
            lines.append(t)

    print("=== DOCX: Training Guide (preview) ===")
    print(f"Path: {docx_path}")
    for i, line in enumerate(lines[:max_lines], start=1):
        print(f"{i:03d}: {line}")
    print(f"\n[DOCX non-empty paragraphs: {len(lines)}]\n")


def extract_xlsx(xlsx_path: Path, *, max_rows: int = 60, max_cols: int = 40) -> None:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(str(xlsx_path), data_only=True)
    print("=== XLSX: Example sheet (preview) ===")
    print(f"Path: {xlsx_path}")
    print("Sheets:", wb.sheetnames)
    for wsname in wb.sheetnames:
        ws = wb[wsname]
        print(f"\n--- Sheet: {wsname} ---")
        rmax = min(ws.max_row or 0, max_rows)
        cmax = min(ws.max_column or 0, max_cols)
        for r in range(1, rmax + 1):
            row = [ws.cell(r, c).value for c in range(1, cmax + 1)]
            if all(v in (None, "") for v in row):
                continue
            while row and row[-1] in (None, ""):
                row.pop()
            print(f"R{r:02d}: {row}")
        print(f"[ws.max_row={ws.max_row}, ws.max_column={ws.max_column}]")
    print()


def main() -> None:
    docx_path = Path(r"c:\Users\black\Downloads\AI automation test\Copy of Auth Data Entry Training Guide.docx")
    xlsx_path = Path(r"c:\Users\black\Downloads\AI automation test\AI automation example sheet 1.xlsx")

    extract_docx(docx_path)
    extract_xlsx(xlsx_path)


if __name__ == "__main__":
    main()

